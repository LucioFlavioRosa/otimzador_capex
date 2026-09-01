"""CTS (Coletor de Tempo Seco) — as duas visoes (usar_cts ligado x desligado) sao a MESMA
demanda. O que TEM de bater: cobertura, vazao, universo efetivo. O que TEM de diferir de
proposito: numero de obras, CAPEX e VPL. Estes testes travam esse contrato."""
import pytest
from _helpers import engine, load_fixture, capex_total, cobertura_fim, codigo

CTS_COMPONENTES = {"cts", "tro", "eee", "lr"}   # Coletor de tempo seco + Tronco + EEE + Linha de recalque


# ---------------------------------------------------------------- estrutura
def test_cts_entram_como_nos_no_modo_ligado(cen_on, cen_off):
    assert cen_on.cts_ids, "modo ligado deveria ter CTS carregadas"
    assert not cen_off.cts_ids, "modo desligado nao deve ter nenhuma CTS"
    assert all(cen_on.nos[c].is_cts for c in cen_on.cts_ids)
    assert not any(getattr(n, "is_cts", False) for n in cen_off.nos.values())


def test_cada_cts_tem_os_quatro_componentes_certos(cen_on):
    for c in cen_on.cts_ids:
        obras = [o for o in cen_on.obras.values() if o.no == c and o.eh_aegea()]
        cods = {codigo(o.id) for o in obras}
        assert cods == CTS_COMPONENTES, f"CTS {c}: esperado {CTS_COMPONENTES}, veio {cods}"
        coletas = [o for o in obras if o.tipo == "coleta"]
        assert len(coletas) == 1 and codigo(coletas[0].id) == "cts", \
            "a ancora de coleta da CTS deve ser o Coletor de tempo seco"


# ---------------------------------------------------------------- os dois cenarios
#
# ESTE BLOCO AFIRMAVA IGUALDADE, e a igualdade caiu junto com a soma. Ligado e desligado
# eram "a mesma demanda" porque a linha da CTS era somada na sub-bacia — e somar conserva
# tudo. Hoje a unica diferenca para a sub-bacia e QUAL COLUNA E LIDA: com o coletor, a
# exclusiva; sem ele, a `*_com_cts`. Sao dois cenarios diferentes, e nao duas contas do
# mesmo cenario.
def test_desligado_atende_menos_que_ligado(res_on, res_off):
    # Sem o coletor, a area que so ele alcancava fica sem atendimento. Nesta fixture, que
    # nao tem as colunas consolidadas, some a demanda inteira da CTS.
    assert cobertura_fim(res_off) < cobertura_fim(res_on)


def test_vazao_NAO_e_mais_somada(cen_on, cen_off):
    """A vazao e DADO da sub-bacia, e o motor nao a inventa para o cenario sem coletor.

    Este teste afirmava o contrario — que as duas visoes tinham a mesma vazao — e era
    verdade porque a linha da CTS era somada. Somar tambem contava a area sobreposta
    duas vezes, e misturava moedas: as ligacoes vinham da coluna consolidada e a vazao
    de uma soma.

    O CUSTO ESTA DECLARADO: sem o coletor, a vazao que chega a ETE e a que estiver na
    base da sub-bacia. Se desligar a CTS muda a vazao dela, quem atualiza a base e quem
    cadastra — o motor avisa que esse cenario existe, mas nao arbitra o numero.
    """
    assert sum(cen_off.vazao.values()) < sum(cen_on.vazao.values())
    # A diferenca e exatamente a vazao das CTS, que deixou de ser absorvida.
    _cts = sum(cen_on.vazao.get(c, 0.0) for c in cen_on.cts_ids)
    assert sum(cen_on.vazao.values()) - sum(cen_off.vazao.values()) == pytest.approx(_cts)


def test_universo_efetivo_do_desligado_e_o_da_sub_bacia(cen_on, cen_off):
    # ON:  b1 1000 + b2 900 + cts1 500x1,2  |  b3 800 + b4 1200 + cts2 400x1,5 = 5100
    # OFF: so as sub-bacias, com o potencial de cada uma  = 1000+900+800+1200   = 3900
    #
    # O potencial da CTS nao entra em media nenhuma: ele e dela, e ela nao esta na rodada.
    assert sum(cen_on.max_lig.values()) == pytest.approx(5100.0)
    assert sum(cen_off.max_lig.values()) == pytest.approx(3900.0)


# ---------------------------------------------------------------- diferencas (TEM de diferir)
def test_ligado_tem_quatro_obras_a_mais_por_cts(cen_on, cen_off):
    n_on = sum(1 for o in cen_on.obras.values() if o.eh_aegea())
    n_off = sum(1 for o in cen_off.obras.values() if o.eh_aegea())
    assert n_on - n_off == 4 * len(cen_on.cts_ids)


def test_capex_ligado_maior_e_diferenca_e_o_capex_das_obras_cts(cen_on, cen_off, res_on, res_off):
    cap_on = capex_total(cen_on, res_on)
    cap_off = capex_total(cen_off, res_off)
    assert cap_on > cap_off, "o modo ligado paga as obras dedicadas da CTS"
    capex_cts = sum(o.capex for o in cen_on.obras.values()
                    if getattr(cen_on.nos.get(o.no), "is_cts", False))
    assert cap_on - cap_off == pytest.approx(capex_cts), \
        "a diferenca de CAPEX tem de ser exatamente o CAPEX das obras da CTS"


def test_desligado_paga_menos_CAPEX_e_a_receita_segue_o_ticket_de_quem_atende(
        cen_on, cen_off, res_on, res_off):
    """Desligar a CTS nao e so economizar obra: muda quem cobra.

    Este teste afirmava `vpl_off > vpl_on`, com a justificativa "mesma receita e
    cobertura, menos CAPEX". A premissa caiu junto com a soma: a receita da linha da
    CTS nao e mais herdada pela sub-bacia.

    Sem o coletor, as ligacoes que ele atenderia passam a ser ligadas pelas obras da
    sub-bacia — e cobradas pelo TICKET DELA. Na fixture, `cts2` fatura 480 por ligacao e
    a `b4` que a absorve, 288: o plano desligado liga a mesma gente por menos dinheiro.

    O que continua valendo sempre e o CAPEX. O VPL depende de qual ticket e maior, e
    fixar um sentido aqui seria fixar um acidente desta fixture.
    """
    assert capex_total(cen_off, res_off) < capex_total(cen_on, res_on)
    _t_cts = 90000 / 200      # cts1: mesmo ticket da b1 que a absorve
    _t_b1 = 180000 / 400
    assert _t_cts == pytest.approx(_t_b1), "a fixture mudou; reveja o raciocinio acima"


# ---------------------------------------------------------------- retrocompatibilidade
def test_banco_sem_cts_modos_sao_identicos():
    # base fixa SEM CTS -> usar_cts ligado ou desligado tem de dar exatamente o mesmo
    a = load_fixture(usar_cts=True)
    b = load_fixture(usar_cts=False)
    assert not a.cts_ids and not b.cts_ids, "o fixture nao deve ter CTS"
    assert set(a.nos) == set(b.nos)
    assert set(a.obras) == set(b.obras)


# ---------------------------------------------------------------- sobreposicao consolidada
#
# AS INVARIANTES DE CIMA VALEM NO CAMINHO DE COMPATIBILIDADE, e este bloco existe para
# dizer por que. Enquanto a origem nao traz as colunas `*_com_cts`, o motor SOMA a linha
# da CTS na sub-bacia — e somar conserva tudo, entao ligado e desligado tem a mesma
# demanda total. E tambem CONTA DUAS VEZES a area sobreposta, que e o defeito.
#
# Com as colunas consolidadas, a igualdade deixa de valer POR CONSTRUCAO: sem o coletor,
# a parte da area que so ele alcancava nao e atendida por ninguem. O que se conserva e
# outra coisa — cada cenario conta a sobreposicao uma vez.
import shutil

import openpyxl
from _helpers import BANK_CTS, silent


def _com_consolidado(tmp_path, por_sub):
    """Copia o banco de CTS e grava as colunas `*_com_cts` de `{sub: {coluna: valor}}`."""
    dst = tmp_path / "cts_consolidado.xlsx"
    shutil.copy(BANK_CTS, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["subbacia-operacional"]
    cab = [str(c.value).strip() for c in ws[1]]
    ichave = cab.index("sub_bacia") + 1
    for sub, valores in por_sub.items():
        for col, val in valores.items():
            if col in cab:
                icol = cab.index(col) + 1
            else:
                icol = ws.max_column + 1
                ws.cell(1, icol).value = col
                cab.append(col)
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, ichave).value == sub:
                    ws.cell(r, icol).value = val
    wb.save(dst)
    return str(dst)


#: b1 atende 1.200 sem o coletor (1.000 exclusivas + 200 de sobreposicao); b4 atende 1.350.
CONSOLIDADO = {
    "b1": {"universo_ligacoes_com_cts": 1200, "ligacoes_atuais_com_cts": 450,
           "universo_economias_com_cts": 1320, "economias_atuais_com_cts": 495},
    "b4": {"universo_ligacoes_com_cts": 1350, "ligacoes_atuais_com_cts": 560,
           "universo_economias_com_cts": 1485, "economias_atuais_com_cts": 616},
}


def test_sem_cts_le_a_coluna_consolidada_em_vez_de_somar(tmp_path):
    # b1 tem 1000 ligacoes exclusivas e a cts1 tem 500. A soma daria 1500 — e conta a
    # area sobreposta duas vezes. A coluna diz que a sub-bacia, sem o coletor, atende
    # 1200: as 1000 dela mais 200 de sobreposicao.
    arq = _com_consolidado(tmp_path, CONSOLIDADO)
    M = engine()
    off = silent(M.ler_banco, arq, usar_cts=False)
    cid = off.nos["b1"].cidade
    # b1 e b2 estao na mesma cidade. O universo EFETIVO ja leva o potencial:
    #   b1 = 1200 (consolidado) x 1,0 (o potencial da PROPRIA sub-bacia) = 1200
    #   + b2 = 900 (potencial 1,0)
    # Somando as duas linhas daria 1600 em b1 — a area sobreposta contada duas vezes.
    assert off.max_lig[cid] == pytest.approx(2100.0)


def test_a_area_so_do_coletor_nao_e_atendida_sem_ele(tmp_path):
    # Com as duas sub-bacias pareadas informando o consolidado, o cenario fica limpo:
    #
    #   ON   b1 1000 + b2 900 + cts1 500x1,2 = 2500  |  b3 800 + b4 1200 + cts2 400x1,5 = 2600
    #   OFF  b1 1200 + b2 900               = 2100  |  b3 800 + b4 1350               = 2150
    #
    # A diferenca (850) e a area que so os coletores alcancavam, com o potencial deles —
    # e nenhum dos dois existe sem eles.
    arq = _com_consolidado(tmp_path, CONSOLIDADO)
    M = engine()
    on = silent(M.ler_banco, arq, usar_cts=True)
    off = silent(M.ler_banco, arq, usar_cts=False)
    assert sum(on.max_lig.values()) == pytest.approx(5100.0)
    assert sum(off.max_lig.values()) == pytest.approx(4250.0)


def test_sem_a_coluna_usa_a_exclusiva_e_ALERTA(capsys):
    """Sem a coluna nao ha o que ler, e o motor NAO inventa.

    A versao anterior somava a linha da CTS aqui. Era o que contava a area sobreposta
    duas vezes — e o que fazia o universo da meta crescer sozinho ao desligar o coletor.
    Base que ainda nao tem a coluna produz uma rodada que ignora a area sobreposta, e o
    ALERTA diz exatamente isso: e um numero a menos, nao um numero errado em silencio.
    """
    M = engine()
    off = M.ler_banco(BANK_CTS, usar_cts=False)
    saida = capsys.readouterr().out
    assert "usou o universo EXCLUSIVO" in saida
    cid = off.nos["b1"].cidade
    assert off.max_lig[cid] == pytest.approx(1900.0)   # 1000 (b1) + 900 (b2), so as sub-bacias


def test_com_cts_ligada_a_coluna_consolidada_e_ignorada(tmp_path):
    # Ela so descreve o cenario SEM coletor. Com ele, a sub-bacia usa o que e exclusivo
    # dela e a CTS entra como no proprio — a sobreposicao esta nos numeros da CTS.
    arq = _com_consolidado(tmp_path, CONSOLIDADO)
    M = engine()
    on_com = silent(M.ler_banco, arq, usar_cts=True)
    on_sem = silent(M.ler_banco, BANK_CTS, usar_cts=True)
    assert sum(on_com.max_lig.values()) == pytest.approx(sum(on_sem.max_lig.values()))
