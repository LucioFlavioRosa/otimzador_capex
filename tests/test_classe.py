"""Recorte da cobertura: total x so residencial (`cobertura_so_residencial`).

O QUE MUDOU, E POR QUE ESTES TESTES MUDARAM JUNTO. O parametro anterior
(`incluir_industrial`) subtraia a parcela industrial de LIGACOES, RECEITA e VAZAO, e
estimava as economias residenciais por proporcao. Ele mexia no VPL: menos receita e
menos vazao mudavam o plano inteiro, e a rodada "so residencial" nao era comparavel
com a outra.

Hoje o recorte acaba na COBERTURA. Quem paga a conta e a ligacao, seja de casa ou de
fabrica — receita, VPL, vazao e CAPEX seguem no total em qualquer modo. O que e
residencial e a META, e ela passou a ser medida com colunas PROPRIAS (`*_residencial`),
vindas do banco, e nao deduzidas por subtracao.

A invariante mais importante desta suite e a de baixo: o VPL NAO PODE MUDAR. Ela e o
que separa esta versao da anterior, e e a que quebraria se alguem voltasse a descontar
industria fora da cobertura.
"""
import shutil

import openpyxl
import pytest
from _helpers import (BANK_CLASSE, BANK_FIXTURE, build_all, capex_total, engine,
                      load_classe, silent)


def _cidade(cen, sb):
    return cen.nos[sb].cidade


# ---------------------------------------------------------------- retrocompat
def test_sem_colunas_residenciais_modos_identicos():
    # Banco antigo, sem `*_residencial`: o recorte nao tem o que recortar e a rodada
    # sai igual. A engine avisa em voz alta (ALERTA) — o que nao pode e mudar numero
    # em silencio.
    M = engine()
    a = silent(M.ler_banco, BANK_FIXTURE, unidade="u1", cobertura_so_residencial=False)
    b = silent(M.ler_banco, BANK_FIXTURE, unidade="u1", cobertura_so_residencial=True)
    assert set(a.obras) == set(b.obras)
    assert sum(a.vazao.values()) == pytest.approx(sum(b.vazao.values()))
    assert sum(a.max_lig.values()) == pytest.approx(sum(b.max_lig.values()))


# ------------------------------------------- o recorte NAO alcanca dinheiro nem obra
def test_capex_igual_entre_os_modos():
    on = load_classe(False); off = load_classe(True)
    assert capex_total(on, build_all(on)) == pytest.approx(capex_total(off, build_all(off)))


def test_vazao_NAO_muda():
    # Vazao dimensiona modulo de ETE e rateia obra compartilhada. Industria contribui
    # com esgoto mesmo quando nao conta para a meta — descontar aqui subdimensionaria a
    # estacao. Era o que a versao anterior fazia.
    on = load_classe(False); off = load_classe(True)
    assert sum(off.vazao.values()) == pytest.approx(sum(on.vazao.values()))


def test_receita_e_VPL_NAO_mudam():
    # A invariante que separa esta versao da anterior. A mesma carteira de obras rende o
    # mesmo, porque a industria continua faturando.
    on = build_all(load_classe(False)); off = build_all(load_classe(True))
    assert sum(off.get("receita_ano", [])) == pytest.approx(sum(on.get("receita_ano", [])))
    assert off["vpl"] == pytest.approx(on["vpl"])


# ---------------------------------------------------------------- cobertura por unidade
def test_universo_da_meta_cai_para_o_residencial():
    # b1: 1000 ligacoes, 800 residenciais; medida em ECONOMIAS (1,1 por ligacao).
    on = load_classe(False); off = load_classe(True)
    c1 = _cidade(on, "b1")
    assert off.max_lig[c1] < on.max_lig[c1] - 1


def test_base_atendida_tambem_cai():
    # Nao basta o denominador virar residencial: a base ja atendida tem de vir da mesma
    # coluna, senao a cobertura de partida mistura as duas moedas e nasce inflada.
    on = load_classe(False); off = load_classe(True)
    c1 = _cidade(on, "b1")
    assert off.base_lig[c1] < on.base_lig[c1] - 1


def test_cobertura_populacao_intacta():
    # Industria nao mora: o universo de populacao ja e residencial, e nao ha coluna
    # `populacao_*_residencial` para existir.
    on = load_classe(False); off = load_classe(True)
    c2 = _cidade(on, "b3")
    assert off.max_lig[c2] == pytest.approx(on.max_lig[c2])


def test_a_obra_conta_menos_para_a_meta_e_o_mesmo_para_a_receita():
    # As duas quantidades da obra, uma ao lado da outra: `lig` (receita) fica; `lig_cob`
    # (meta) encolhe. Antes eram o mesmo campo, e era por isso que o recorte nao cabia.
    off = load_classe(True)
    coletas = [o for o in off.coletas if o.no == "b1"]
    assert coletas, "a fixture precisa ter coleta em b1"
    o = coletas[0]
    assert o.lig_cob < o.lig
    assert o.lig == pytest.approx(600.0)      # 1000 - 400, total
    assert o.lig_cob == pytest.approx(480.0)  # 800 - 320, residencial


def test_sem_recorte_as_duas_quantidades_sao_iguais():
    on = load_classe(False)
    for o in on.coletas:
        assert o.lig_cob == pytest.approx(o.lig)


def test_cobertura_em_ligacoes_tambem_cai(tmp_path):
    # muda c1 para LIGACOES: o recorte nao pode depender da unidade de cobertura.
    dst = tmp_path / "classe_ligacoes.xlsx"
    shutil.copy(BANK_CLASSE, dst)
    wb = openpyxl.load_workbook(dst); ws = wb["cidade-operacional"]
    h = [c.value for c in ws[1]]; ic = h.index("cidade_id") + 1; iu = h.index("unidade_cobertura") + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, ic).value == "c1":
            ws.cell(r, iu).value = "ligacoes"
    wb.save(dst)
    M = engine()
    on = silent(M.ler_banco, str(dst), cobertura_so_residencial=False)
    off = silent(M.ler_banco, str(dst), cobertura_so_residencial=True)
    c1 = on.nos["b1"].cidade
    assert off.max_lig[c1] < on.max_lig[c1] - 1
    # O universo e AGREGADO POR CIDADE: c1 tem b1 (1000 ligacoes, 800 residenciais) e b2
    # (900, sem industria, logo 900 residenciais). 800 + 900 = 1700, sem conversao de
    # unidade porque aqui a cidade mede em ligacoes.
    assert off.max_lig[c1] == pytest.approx(1700.0)
    assert on.max_lig[c1] == pytest.approx(1900.0)
