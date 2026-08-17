"""Colunas DERIVADAS pela engine (nao sao input do usuario):
  - *_novas_obras = max(0, universo - atuais);
  - a engine IGNORA o valor do banco e usa o derivado (com aviso na divergencia).
"""
import shutil
import openpyxl
import pytest
from _helpers import engine, silent, BANK_CLASSE


def _subop(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["subbacia-operacional"]; it = ws.iter_rows(values_only=True); h = list(next(it))
    return {d["sub_bacia"]: d for d in (dict(zip(h, r)) for r in it)}


def test_ligacoes_novas_e_universo_menos_atuais():
    # coleta.lig (ligacoes novas das obras) deve ser universo_ligacoes - ligacoes_atuais
    M = engine()
    so = _subop(BANK_CLASSE)
    cen = silent(M.ler_banco, BANK_CLASSE, cobertura_so_residencial=False)
    checados = 0
    for col in cen.coletas:
        d = so.get(col.no)
        if not d:
            continue
        esperado = max(0.0, float(d["universo_ligacoes"]) - float(d["ligacoes_atuais"]))
        assert col.lig == pytest.approx(esperado), f"{col.no}: {col.lig} != {esperado}"
        checados += 1
    assert checados > 0


def test_valor_do_banco_e_ignorado(tmp_path):
    # zera ligacoes_novas_obras no banco; a engine deve DERIVAR (universo-atuais), nao usar o 0
    M = engine()
    dst = tmp_path / "novas_zerado.xlsx"
    shutil.copy(BANK_CLASSE, dst)
    wb = openpyxl.load_workbook(dst); ws = wb["subbacia-operacional"]
    h = [c.value for c in ws[1]]; idx = {x: i + 1 for i, x in enumerate(h)}
    for r in range(2, ws.max_row + 1):
        ws.cell(r, idx["ligacoes_novas_obras"]).value = 0
    wb.save(dst)
    cen = silent(M.ler_banco, str(dst), cobertura_so_residencial=False)
    # se a engine tivesse usado o 0 do banco, nenhuma coleta teria ligacoes novas
    assert any(col.lig > 0 for col in cen.coletas), "a engine usou o valor (errado) do banco em vez de derivar"
